"""
Admin blueprint for dashboard, employees, attendance, reports, and settings
"""
from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import date, datetime, timedelta
from utils.logger import get_logger
from utils.validators import (
    validate_required, validate_employee_registration_data, 
    validate_integer, validate_float, validate_date, validate_time,
    ValidationError, sanitize_string
)
from utils.security import hash_password
import sqlite3
import json
import os
import base64
from io import BytesIO
from PIL import Image
import face_utils

logger = get_logger(__name__)
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def _ensure_proof_file_path_column(db_cursor):
    """Ensure attendance table has proof_file_path column (migration for existing DBs)."""
    try:
        db_cursor.execute("PRAGMA table_info(attendance)")
        columns = [row[1] for row in db_cursor.fetchall()]
        if "proof_file_path" not in columns:
            db_cursor.execute("ALTER TABLE attendance ADD COLUMN proof_file_path TEXT")
    except Exception:
        pass


def _ensure_audit_log_table(db_cursor):
    """Ensure audit_log table exists (for tracking sensitive admin actions)."""
    try:
        db_cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_name TEXT,
                action TEXT NOT NULL,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    except Exception:
        # Logging is best-effort; do not break main flows
        logger.warning("Could not ensure audit_log table exists", exc_info=True)


def _ensure_admin_photo_column(db_cursor):
    """Ensure admin table has photo_path column for profile photo."""
    try:
        db_cursor.execute("PRAGMA table_info(admin)")
        columns = [row[1] for row in db_cursor.fetchall()]
        if "photo_path" not in columns:
            db_cursor.execute("ALTER TABLE admin ADD COLUMN photo_path TEXT")
    except Exception:
        pass


def log_admin_action(db, action: str, details: str = "", admin_name: str | None = None):
    """
    Record an admin action into audit_log.
    Safe to call inside request handlers; failures are logged but ignored.
    """
    try:
        cur = db.cursor()
        _ensure_audit_log_table(cur)
        if admin_name is None:
            admin_name = session.get("admin_name")
        cur.execute(
            "INSERT INTO audit_log (admin_name, action, details) VALUES (?, ?, ?)",
            (admin_name, action, details[:1000]),
        )
        db.commit()
    except Exception as exc:
        logger.warning(f"Failed to write audit log entry for action '{action}': {exc}", exc_info=True)


def admin_required(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin_logged_in"):
            logger.warning("Unauthorized access attempt to admin route")
            return redirect(url_for("auth.admin_login"))
        return f(*args, **kwargs)
    return wrapper


@admin_bp.route("/dashboard")
@admin_required
def dashboard():
    """Admin dashboard with statistics"""
    try:
        from utils.helpers import ensure_absent_records_for_date
        from db import get_db
        db = get_db()
        cur = db.cursor()
        today = date.today().isoformat()

        # Auto-create Absent records for employees with no attendance today
        ensure_absent_records_for_date(today)

        cur.execute("""
            SELECT 
                (SELECT COUNT(*) FROM employees) as total_employees,
                COALESCE(SUM(CASE WHEN date = ? AND (attendance_status = 'Present' OR attendance_status = 'Late') THEN 1 ELSE 0 END), 0) as present,
                COALESCE(SUM(CASE WHEN date = ? AND attendance_status = 'Absent' THEN 1 ELSE 0 END), 0) as absent,
                COALESCE(SUM(CASE WHEN date = ? AND attendance_status = 'Late' THEN 1 ELSE 0 END), 0) as late
            FROM attendance
        """, (today, today, today))
        
        stats = cur.fetchone()
        total_employees = stats[0] or 0
        present = stats[1] or 0
        absent = stats[2] or 0
        late = stats[3] or 0

        cur.execute("""
            SELECT e.full_name, e.employee_code, a.date, a.morning_in, a.lunch_out,
                   a.afternoon_in, a.time_out, a.attendance_status, a.verification_method
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            ORDER BY a.attendance_id DESC
            LIMIT 10
        """)
        recent_attendance = cur.fetchall()

        logger.info(f"Dashboard accessed by admin: {session.get('admin_name')}")
        return render_template(
            "admin/dashboard.html",
            total_employees=total_employees,
            present=present,
            absent=absent,
            late=late,
            recent_attendance=recent_attendance,
        )
    except Exception as e:
        logger.error(f"Dashboard error: {str(e)}", exc_info=True)
        return render_template("admin/dashboard.html", 
                             total_employees=0, present=0, absent=0, late=0, 
                             recent_attendance=[], error="Error loading dashboard")


@admin_bp.route("/register", methods=["GET", "POST"])
@admin_required
def register():
    """Register new employee"""
    if request.method == "POST":
        try:
            from app import get_db
            db = get_db()
            cur = db.cursor()

            # Validate all input data
            form_data = dict(request.form)
            validated_data, errors = validate_employee_registration_data(form_data)
            
            if errors:
                logger.warning(f"Employee registration validation errors: {errors}")
                return render_template("admin/register.html", error="; ".join(errors))

            # Handle photo saving
            photo_path = None
            face_images = validated_data.get('face_images', [])
            if face_images and len(face_images) > 0 and face_images[0]:
                try:
                    img_data = face_images[0]
                    if ',' in img_data:
                        img_data = img_data.split(',')[1]
                    
                    image_data = base64.b64decode(img_data)
                    image = Image.open(BytesIO(image_data))
                    
                    photos_dir = os.path.join('static', 'photos')
                    os.makedirs(photos_dir, exist_ok=True)
                    
                    photo_filename = f"employee_{validated_data['employee_code']}_{validated_data['full_name'].replace(' ', '_')}.jpg"
                    photo_path = os.path.join(photos_dir, photo_filename)
                    image.save(photo_path, 'JPEG', quality=85)
                    photo_path = f"photos/{photo_filename}"
                    logger.info(f"Photo saved for employee: {photo_path}")
                except Exception as e:
                    logger.warning(f"Error saving photo: {str(e)}")

            # Insert employee
            cur.execute("""
                INSERT INTO employees (
                    full_name, employee_code, address, place_of_birth, blood_type,
                    date_of_birth, gender, civil_status, age,
                    contact_number, email, course, entity_office,
                    bp_number, philhealth_number, pagibig_number, tin, id_number,
                    position, salary_grade, basic_salary, department,
                    place_of_assignment, original_place_of_assignment, item_number,
                    date_appointed, date_of_last_promotion, date_of_separation,
                    employment_status, eligibility, photo_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                validated_data['full_name'],
                validated_data['employee_code'],
                validated_data.get('address', ''),
                validated_data.get('place_of_birth', ''),
                validated_data.get('blood_type', ''),
                validated_data.get('date_of_birth'),
                validated_data.get('gender', ''),
                validated_data.get('civil_status', ''),
                validated_data.get('age'),
                validated_data.get('contact_number', ''),
                validated_data.get('email', ''),
                validated_data.get('course', ''),
                validated_data.get('entity_office', ''),
                validated_data.get('bp_number', ''),
                validated_data.get('philhealth_number', ''),
                validated_data.get('pagibig_number', ''),
                validated_data.get('tin', ''),
                validated_data.get('id_number', ''),
                validated_data.get('position', ''),
                validated_data.get('salary_grade', ''),
                validated_data.get('basic_salary'),
                validated_data.get('department', ''),
                validated_data.get('place_of_assignment', ''),
                validated_data.get('original_place_of_assignment', ''),
                validated_data.get('item_number', ''),
                validated_data.get('date_appointed'),
                validated_data.get('date_of_last_promotion'),
                validated_data.get('date_of_separation'),
                validated_data.get('employment_status', ''),
                validated_data.get('eligibility', ''),
                photo_path,
            ))

            employee_id = cur.lastrowid
            db.commit()
            logger.info(f"Employee registered successfully: ID={employee_id}, Name={validated_data['full_name']}, Code={validated_data['employee_code']}")

            # Handle face encodings
            if face_images:
                try:
                    for img_data in face_images:
                        encoding = face_utils.encode_face_from_base64(img_data)
                        if encoding is not None:
                            face_utils.save_face(employee_id, encoding)
                            logger.info(f"Face encoding saved for employee {employee_id}")
                            break
                except Exception as e:
                    logger.warning(f"Error processing face images: {str(e)}")

            return redirect(url_for("admin.employees"))

        except sqlite3.IntegrityError as e:
            db.rollback()
            error_msg = f"Employee ID '{validated_data.get('employee_code', '')}' already exists"
            logger.error(f"Registration failed - IntegrityError: {error_msg}")
            return render_template("admin/register.html", error=error_msg)
        except Exception as e:
            db.rollback()
            error_msg = f"Error: {str(e)}"
            logger.error(f"Registration error: {error_msg}", exc_info=True)
            return render_template("admin/register.html", error=error_msg)

    return render_template("admin/register.html")


@admin_bp.route("/employees")
@admin_required
def employees():
    """List all employees"""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        search_query = request.args.get("search", "").strip()
        
        if search_query:
            search_pattern = f"%{search_query}%"
            cur.execute("""
                SELECT * FROM employees 
                WHERE full_name LIKE ? OR employee_code LIKE ? OR department LIKE ?
                ORDER BY full_name
            """, (search_pattern, search_pattern, search_pattern))
        else:
            cur.execute("SELECT * FROM employees ORDER BY full_name")
        
        employees = cur.fetchall()
        logger.debug(f"Employees list accessed, found {len(employees)} employees")
        return render_template("admin/employees.html", employees=employees, search_query=search_query)
    except Exception as e:
        logger.error(f"Error in employees route: {str(e)}", exc_info=True)
        return render_template("admin/employees.html", employees=[], search_query=search_query or "")


@admin_bp.route("/employee-info")
@admin_required
def employee_info():
    """Employee information page"""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        employee_id = request.args.get("id")
        
        if employee_id:
            # Try to convert to int for safety
            try:
                employee_id = int(employee_id)
            except (ValueError, TypeError):
                logger.warning(f"Invalid employee_id provided: {request.args.get('id')}")
                employee_id = None
            
            if employee_id:
                cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
                employee = cur.fetchone()
                if employee:
                    # sqlite3.Row objects use bracket notation
                    employee_name = employee["full_name"] if "full_name" in employee.keys() else "N/A"
                    logger.info(f"Employee found: ID={employee_id}, Name={employee_name}")
                    return render_template("admin/employee_info.html", employee=employee, employees=[])
                else:
                    logger.warning(f"Employee not found for ID: {employee_id}")
                    # Employee not found - show list with error message
                    cur.execute("SELECT id, full_name, employee_code, department FROM employees ORDER BY full_name")
                    employees = cur.fetchall()
                    return render_template("admin/employee_info.html", employee=None, employees=employees, error=f"Employee with ID {employee_id} not found.")
        
        # No employee_id provided - show list
        cur.execute("SELECT id, full_name, employee_code, department FROM employees ORDER BY full_name")
        employees = cur.fetchall()
        logger.debug(f"Showing employee list with {len(employees)} employees")
        return render_template("admin/employee_info.html", employee=None, employees=employees)
    except Exception as e:
        logger.error(f"Error in employee_info route: {str(e)}", exc_info=True)
        # On error, try to show the list anyway
        try:
            from db import get_db
            db = get_db()
            cur = db.cursor()
            cur.execute("SELECT id, full_name, employee_code, department FROM employees ORDER BY full_name")
            employees = cur.fetchall()
            return render_template("admin/employee_info.html", employee=None, employees=employees, error=f"Error loading employee information: {str(e)}")
        except:
            return render_template("admin/employee_info.html", employee=None, employees=[], error="Error loading employee information. Please try again.")


@admin_bp.route("/employee/delete/<int:employee_id>", methods=["POST"])
@admin_required
def delete_employee(employee_id):
    """Delete an employee"""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        
        # Get employee info before deletion
        cur.execute("SELECT photo_path, full_name FROM employees WHERE id=?", (employee_id,))
        employee = cur.fetchone()
        
        if not employee:
            logger.warning(f"Employee {employee_id} not found for deletion")
            return redirect(url_for("admin.employees"))
        
        employee_name = employee["full_name"] if "full_name" in employee.keys() else "Unknown"
        
        # Delete photo file if exists
        # sqlite3.Row objects use bracket notation, not .get()
        if employee and "photo_path" in employee.keys() and employee["photo_path"]:
            try:
                photo_path = os.path.join("static", employee["photo_path"])
                if os.path.exists(photo_path):
                    os.remove(photo_path)
                    logger.info(f"Photo deleted: {photo_path}")
            except Exception as e:
                logger.warning(f"Could not delete photo file: {str(e)}")
        
        # IMPORTANT: Delete related data FIRST to avoid foreign key constraint errors
        # Delete in reverse order of dependencies
        logger.info(f"Deleting related data for employee {employee_id} ({employee_name})")
        
        # 1. Delete attendance records first
        cur.execute("DELETE FROM attendance WHERE employee_id=?", (employee_id,))
        attendance_deleted = cur.rowcount
        logger.info(f"Deleted {attendance_deleted} attendance records")
        
        # 2. Delete facial data
        cur.execute("DELETE FROM facial_data WHERE employee_id=?", (employee_id,))
        facial_deleted = cur.rowcount
        logger.info(f"Deleted {facial_deleted} facial data records")
        
        # 3. Finally delete the employee record
        cur.execute("DELETE FROM employees WHERE id=?", (employee_id,))
        
        db.commit()
        logger.info(f"Employee {employee_id} ({employee_name}) deleted successfully")
        return redirect(url_for("admin.employees"))
    except Exception as e:
        try:
            db.rollback()
        except:
            pass
        logger.error(f"Error deleting employee {employee_id}: {str(e)}", exc_info=True)
        return redirect(url_for("admin.employee_info", id=employee_id, error=f"Failed to delete employee: {str(e)}"))


@admin_bp.route("/employee/edit/<int:employee_id>", methods=["GET", "POST"])
@admin_required
def edit_employee(employee_id):
    """Edit employee information"""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        
        if request.method == "POST":
            # Validate input
            full_name = validate_required(request.form.get("full_name", "").strip(), "Full Name")
            
            # Convert optional fields
            age = validate_integer(request.form.get("age", ""), "Age", min_value=16, max_value=100)
            basic_salary = validate_float(request.form.get("basic_salary", ""), "Basic Salary", min_value=0)
            
            # Handle photo update
            photo_path = None
            face_images = request.form.get("face_images", "[]")
            if face_images and face_images != "[]":
                try:
                    images = json.loads(face_images)
                    if images and images[0]:
                        img_data = images[0]
                        if ',' in img_data:
                            img_data = img_data.split(',')[1]
                        
                        image_data = base64.b64decode(img_data)
                        image = Image.open(BytesIO(image_data))
                        
                        photos_dir = os.path.join('static', 'photos')
                        os.makedirs(photos_dir, exist_ok=True)
                        
                        cur.execute("SELECT employee_code, full_name FROM employees WHERE id=?", (employee_id,))
                        emp_data = cur.fetchone()
                        emp_code = emp_data["employee_code"] if emp_data else str(employee_id)
                        emp_name = full_name
                        
                        photo_filename = f"employee_{emp_code}_{emp_name.replace(' ', '_')}.jpg"
                        photo_path = os.path.join(photos_dir, photo_filename)
                        image.save(photo_path, 'JPEG', quality=85)
                        photo_path = f"photos/{photo_filename}"
                        logger.info(f"Photo updated for employee {employee_id}")
                except Exception as e:
                    logger.warning(f"Error saving photo: {str(e)}")
            
            # Update employee
            update_fields = [
                'full_name', 'address', 'place_of_birth', 'blood_type',
                'date_of_birth', 'gender', 'civil_status', 'age',
                'contact_number', 'email', 'course', 'entity_office',
                'bp_number', 'philhealth_number', 'pagibig_number', 'tin', 'id_number',
                'position', 'salary_grade', 'basic_salary', 'department',
                'place_of_assignment', 'original_place_of_assignment', 'item_number',
                'date_appointed', 'date_of_last_promotion', 'date_of_separation',
                'employment_status', 'eligibility'
            ]
            
            values = [
                full_name,
                sanitize_string(request.form.get("address", "")),
                sanitize_string(request.form.get("place_of_birth", "")),
                sanitize_string(request.form.get("blood_type", "")),
                validate_date(request.form.get("date_of_birth", "")),
                sanitize_string(request.form.get("gender", "")),
                sanitize_string(request.form.get("civil_status", "")),
                age,
                sanitize_string(request.form.get("contact_number", "")),
                sanitize_string(request.form.get("email", "")),
                sanitize_string(request.form.get("course", "")),
                sanitize_string(request.form.get("entity_office", "")),
                sanitize_string(request.form.get("bp_number", "")),
                sanitize_string(request.form.get("philhealth_number", "")),
                sanitize_string(request.form.get("pagibig_number", "")),
                sanitize_string(request.form.get("tin", "")),
                sanitize_string(request.form.get("id_number", "")),
                sanitize_string(request.form.get("position", "")),
                sanitize_string(request.form.get("salary_grade", "")),
                basic_salary,
                sanitize_string(request.form.get("department", "")),
                sanitize_string(request.form.get("place_of_assignment", "")),
                sanitize_string(request.form.get("original_place_of_assignment", "")),
                sanitize_string(request.form.get("item_number", "")),
                validate_date(request.form.get("date_appointed", "")),
                validate_date(request.form.get("date_of_last_promotion", "")),
                validate_date(request.form.get("date_of_separation", "")),
                sanitize_string(request.form.get("employment_status", "")),
                sanitize_string(request.form.get("eligibility", ""))
            ]
            
            if photo_path:
                update_fields.append('photo_path')
                values.append(photo_path)
            
            values.append(employee_id)
            
            placeholders = ', '.join([f"{field}=?" for field in update_fields])
            cur.execute(f"UPDATE employees SET {placeholders} WHERE id=?", values)
            
            db.commit()
            logger.info(f"Employee {employee_id} updated successfully")
            return redirect(url_for("admin.employee_info", id=employee_id))
        
        # GET request - show edit form
        cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
        employee = cur.fetchone()
        if not employee:
            return redirect(url_for("admin.employees"))
        
        return render_template("admin/edit_employee.html", employee=employee)
        
    except ValidationError as e:
        logger.warning(f"Validation error editing employee {employee_id}: {str(e)}")
        cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
        employee = cur.fetchone()
        return render_template("admin/edit_employee.html", employee=employee, error=str(e))
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating employee {employee_id}: {str(e)}", exc_info=True)
        cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
        employee = cur.fetchone()
        return render_template("admin/edit_employee.html", employee=employee, error=str(e))


@admin_bp.route("/attendance")
@admin_required
def attendance():
    """Attendance management page"""
    try:
        from utils.helpers import ensure_absent_records_for_date
        from db import get_db
        db = get_db()
        cur = db.cursor()
        
        selected_date = request.args.get("date", date.today().isoformat())

        # Auto-create Absent records for employees with no attendance on the selected date
        ensure_absent_records_for_date(selected_date)
        _ensure_proof_file_path_column(cur)
        cur.execute("""
            SELECT a.attendance_id, e.id, e.full_name, a.date, a.morning_in, a.lunch_out,
                   a.afternoon_in, a.time_out, a.attendance_status, a.verification_method, a.proof_file_path
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.date=?
            ORDER BY e.full_name
        """, (selected_date,))

        records = cur.fetchall()
        return render_template(
            "admin/attendance.html",
            records=records,
            selected_date=selected_date,
        )
    except Exception as e:
        logger.error(f"Error in attendance route: {str(e)}", exc_info=True)
        return render_template("admin/attendance.html", records=[], selected_date=date.today().isoformat())


@admin_bp.route("/attendance/edit/<int:attendance_id>", methods=["GET", "POST"])
@admin_required
def edit_dtr(attendance_id):
    """Edit Daily Time Record"""
    try:
        from db import get_db
        from utils.helpers import check_if_late
        db = get_db()
        cur = db.cursor()
        
        if request.method == "POST":
            # Get form data
            if request.is_json:
                data = request.get_json()
                morning_in = data.get("morning_in", "").strip() or None
                lunch_out = data.get("lunch_out", "").strip() or None
                afternoon_in = data.get("afternoon_in", "").strip() or None
                time_out = data.get("time_out", "").strip() or None
                attendance_date = validate_required(data.get("date", ""), "Date")
            else:
                morning_in = validate_time(request.form.get("morning_in", ""), "Morning In") if request.form.get("morning_in") else None
                lunch_out = validate_time(request.form.get("lunch_out", ""), "Lunch Out") if request.form.get("lunch_out") else None
                afternoon_in = validate_time(request.form.get("afternoon_in", ""), "Afternoon In") if request.form.get("afternoon_in") else None
                time_out = validate_time(request.form.get("time_out", ""), "Time Out") if request.form.get("time_out") else None
                attendance_date = validate_required(request.form.get("date", ""), "Date")
            
            # Get existing record (to enforce edit window)
            _ensure_proof_file_path_column(cur)
            cur.execute("SELECT * FROM attendance WHERE attendance_id=?", (attendance_id,))
            existing = cur.fetchone()
            
            if not existing:
                if request.is_json:
                    return jsonify({"success": False, "message": "Attendance record not found"}), 404
                return redirect(url_for("admin.attendance"))
            
            # Proof file upload (travel order, memo, etc.) - form only
            proof_file_path = None
            if not request.is_json:
                from config import PROOF_UPLOADS_DIR, ALLOWED_PROOF_EXTENSIONS, MAX_PROOF_FILE_SIZE
                os.makedirs(PROOF_UPLOADS_DIR, exist_ok=True)
                try:
                    existing_path = existing["proof_file_path"]
                except (KeyError, TypeError):
                    existing_path = None
                proof_file_path = existing_path
                f = request.files.get("proof_file")
                if f and f.filename:
                    ext = (os.path.splitext(secure_filename(f.filename))[1] or "").lstrip(".").lower()
                    if ext not in ALLOWED_PROOF_EXTENSIONS:
                        record = dict(existing) if existing else {}
                        return render_template("admin/edit_dtr.html", record=record, error="Invalid file type. Allowed: PDF, JPG, PNG, DOC, DOCX")
                    f.seek(0, 2)
                    size = f.tell()
                    f.seek(0)
                    if size > MAX_PROOF_FILE_SIZE:
                        record = dict(existing) if existing else {}
                        return render_template("admin/edit_dtr.html", record=record, error="Proof file is too large (max 10MB)")
                    from datetime import datetime
                    safe_name = f"att_{attendance_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                    save_path = os.path.join(PROOF_UPLOADS_DIR, safe_name)
                    f.save(save_path)
                    proof_file_path = os.path.join("uploads", "proofs", safe_name).replace("\\", "/")
            
            # Proof/remarks when editing (e.g. employee submitted proof for forgotten attendance)
            proof_notes = None
            if request.is_json:
                proof_notes = (request.get_json() or {}).get("proof_notes", "").strip() or None
                try:
                    proof_file_path = existing["proof_file_path"]
                except (KeyError, TypeError):
                    proof_file_path = None
            else:
                proof_notes = (request.form.get("proof_notes") or "").strip() or None
            
            # Determine attendance status
            attendance_status = "Present"
            if morning_in:
                try:
                    if check_if_late(morning_in, "morning"):
                        attendance_status = "Late"
                except:
                    pass
            
            if afternoon_in:
                try:
                    if attendance_status != "Late" and check_if_late(afternoon_in, "afternoon"):
                        attendance_status = "Late"
                except:
                    pass
            
            # Set verification_method: "Admin edit" or "Admin edit; Proof: ..."
            verification = "Admin edit"
            if proof_notes:
                from utils.validators import sanitize_string
                verification = "Admin edit; Proof: " + sanitize_string(proof_notes, max_length=400)
            
            # Update attendance record (include verification_method and proof file path)
            cur.execute("""
                UPDATE attendance 
                SET morning_in=?, lunch_out=?, afternoon_in=?, time_out=?, 
                    attendance_status=?, date=?, verification_method=?, proof_file_path=?
                WHERE attendance_id=?
            """, (morning_in, lunch_out, afternoon_in, time_out, attendance_status, attendance_date, verification, proof_file_path, attendance_id))
            
            db.commit()
            logger.info(f"DTR {attendance_id} updated successfully")
            log_admin_action(
                db,
                action="edit_dtr",
                details=f"Edited DTR {attendance_id} for date {attendance_date}",
            )
            
            if request.is_json:
                return jsonify({"success": True, "message": "DTR updated successfully"})
            return redirect(url_for("admin.attendance", date=attendance_date))
        
        # GET request
        _ensure_proof_file_path_column(cur)
        cur.execute("""
            SELECT a.*, e.full_name, e.employee_code
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.attendance_id=?
        """, (attendance_id,))
        
        row = cur.fetchone()
        if not row:
            return redirect(url_for("admin.attendance"))
        # Convert sqlite3.Row to dict so template/filters never call .get() on Row
        record = dict(row) if row else None
        return render_template("admin/edit_dtr.html", record=record)
        
    except ValidationError as e:
        logger.warning(f"Validation error editing DTR {attendance_id}: {str(e)}")
        if request.is_json:
            return jsonify({"success": False, "message": str(e)}), 400
        cur.execute("SELECT a.*, e.full_name, e.employee_code FROM attendance a JOIN employees e ON a.employee_id = e.id WHERE a.attendance_id=?", (attendance_id,))
        row = cur.fetchone()
        record = dict(row) if row else None
        return render_template("admin/edit_dtr.html", record=record, error=str(e))
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating DTR {attendance_id}: {str(e)}", exc_info=True)
        if request.is_json:
            return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
        cur.execute("SELECT a.*, e.full_name, e.employee_code FROM attendance a JOIN employees e ON a.employee_id = e.id WHERE a.attendance_id=?", (attendance_id,))
        row = cur.fetchone()
        record = dict(row) if row else None
        return render_template("admin/edit_dtr.html", record=record, error=str(e))


@admin_bp.route("/audit-log")
@admin_required
def audit_log():
    """Read-only view of recent audit log entries."""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        _ensure_audit_log_table(cur)
        cur.execute(
            """
            SELECT id, admin_name, action, details, created_at
            FROM audit_log
            ORDER BY id DESC
            LIMIT 200
            """
        )
        logs = cur.fetchall()
        return render_template("admin/audit_log.html", logs=logs)
    except Exception as e:
        logger.error(f"Error loading audit log: {str(e)}", exc_info=True)
        return render_template("admin/audit_log.html", logs=[], error=str(e))


@admin_bp.route("/audit-log/delete/<int:log_id>", methods=["POST"])
@admin_required
def delete_audit_log(log_id):
    """Delete a single audit log entry (admin only)."""
    from flask import jsonify
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        cur.execute("DELETE FROM audit_log WHERE id = ?", (log_id,))
        deleted = cur.rowcount
        db.commit()
        if deleted:
            logger.info(f"Audit log entry {log_id} deleted by admin")
            return jsonify({"success": True, "message": "Log entry deleted."})
        return jsonify({"success": False, "error": "Log entry not found."}), 404
    except Exception as e:
        logger.error(f"Error deleting audit log {log_id}: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route("/reports")
@admin_required
def reports():
    """Reports/DTR generation page"""
    try:
        from db import get_db
        db = get_db()
        cur = db.cursor()
        
        cur.execute("SELECT id, full_name, employee_code FROM employees ORDER BY full_name")
        employees = cur.fetchall()
        
        employee_id = request.args.get("employee_id")
        month_year = request.args.get("month", "")
        
        dtr_data = None
        employee_info = None
        
        if employee_id and month_year:
            cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
            employee_info = cur.fetchone()
            
            try:
                year, month = month_year.split("-")
                cur.execute("""
                    SELECT date, morning_in, lunch_out, afternoon_in, time_out, attendance_status
                    FROM attendance
                    WHERE employee_id=? AND date LIKE ?
                    ORDER BY date
                """, (employee_id, f"{year}-{month}-%"))
                
                attendance_records = cur.fetchall()
                
                dtr_data = {
                    "year": year,
                    "month": month,
                    "records": {},
                    "total_work_minutes": 0,
                }

                def _compute_work_minutes(morning_in, lunch_out, afternoon_in, time_out):
                    """
                    Compute total work minutes for the day.
                    - Count morning segment only if BOTH morning_in AND lunch_out exist (complete morning half).
                    - Count afternoon segment only if BOTH afternoon_in AND time_out exist (complete afternoon half).
                    - Half-day is OK: only morning or only afternoon still gets totalled.
                    - Do NOT total if a segment has only one punch (e.g. morning in but no lunch out).
                    - Lunch break (12:00 PM - 1:00 PM) not counted; afternoon count starts at 1:00 PM if earlier.
                    """
                    try:
                        fmt = "%I:%M %p"
                        lunch_start = datetime.strptime("12:00 PM", fmt)
                        afternoon_sched_start = datetime.strptime("01:00 PM", fmt)
                        morning_minutes = 0
                        afternoon_minutes = 0

                        if morning_in and lunch_out:
                            m_in = datetime.strptime(morning_in, fmt)
                            l_out = datetime.strptime(lunch_out, fmt)
                            morning_end = min(l_out, lunch_start)
                            morning_minutes = max(
                                0, int((morning_end - m_in).total_seconds() / 60)
                            )

                        if afternoon_in and time_out:
                            a_in = datetime.strptime(afternoon_in, fmt)
                            t_out = datetime.strptime(time_out, fmt)
                            afternoon_start = max(a_in, afternoon_sched_start)
                            afternoon_minutes = max(
                                0, int((t_out - afternoon_start).total_seconds() / 60)
                            )

                        work_minutes = morning_minutes + afternoon_minutes
                        return work_minutes if work_minutes > 0 else 0
                    except Exception:
                        return 0
                
                for record in attendance_records:
                    day = record["date"].split("-")[2]
                    morning_in = record["morning_in"] or ""
                    lunch_out = record["lunch_out"] or ""
                    afternoon_in = record["afternoon_in"] or ""
                    time_out = record["time_out"] or ""
                    work_minutes = _compute_work_minutes(
                        morning_in, lunch_out, afternoon_in, time_out
                    )
                    dtr_data["records"][day] = {
                        "morning_in": morning_in,
                        "lunch_out": lunch_out,
                        "afternoon_in": afternoon_in,
                        "time_out": time_out,
                        "status": record["attendance_status"] or "",
                        "work_minutes": work_minutes,
                    }
                    dtr_data["total_work_minutes"] += work_minutes
            except Exception as e:
                logger.warning(f"Error processing DTR data: {str(e)}")
        
        return render_template(
            "admin/reports.html",
            employees=employees,
            employee_id=employee_id,
            month_year=month_year,
            dtr_data=dtr_data,
            employee_info=employee_info
        )
    except Exception as e:
        logger.error(f"Error in reports route: {str(e)}", exc_info=True)
        return render_template("admin/reports.html", employees=[], dtr_data=None, employee_info=None)


@admin_bp.route("/settings", methods=["GET", "POST"])
@admin_required
def settings():
    """Settings page for holidays and suspensions"""
    db = None
    try:
        from db import get_db
        from utils.helpers import get_time_settings, clear_time_settings_cache
        from datetime import datetime
        
        # Load time settings
        time_settings_raw = get_time_settings()
        
        # Convert time format from "HH:MM AM/PM" to "HH:MM" for HTML time inputs
        def convert_time_for_html(time_str):
            """Convert '06:00 AM' to '06:00' for HTML time input"""
            try:
                if time_str and ('AM' in time_str or 'PM' in time_str):
                    time_obj = datetime.strptime(time_str, "%I:%M %p")
                    return time_obj.strftime("%H:%M")
                return time_str or ""
            except:
                return time_str or ""
        
        time_settings = {}
        for key, value in time_settings_raw.items():
            time_settings[key] = convert_time_for_html(value)
        
        db = get_db()
        cur = db.cursor()
        _ensure_admin_photo_column(cur)
        current_admin = None
        if session.get("admin_id"):
            cur.execute("SELECT id, name, photo_path FROM admin WHERE id = ?", (session["admin_id"],))
            row = cur.fetchone()
            if row:
                current_admin = {"id": row["id"], "name": row["name"] or "", "photo_path": (row["photo_path"] if "photo_path" in row.keys() else None) or ""}
        
        if request.method == "POST":
            action = request.form.get("action")
            
            if action == "save_time_settings":
                # Save attendance time settings
                from utils.validators import ValidationError, validate_required

                def parse_html_time(html_time_str, field_label):
                    """Convert 'HH:MM' (24h) from HTML input to 'HH:MM AM/PM'."""
                    try:
                        value = validate_required(html_time_str, field_label)
                        t = datetime.strptime(value, "%H:%M")
                        return t.strftime("%I:%M %p")
                    except ValidationError:
                        raise
                    except Exception:
                        raise ValidationError(f"{field_label} must be a valid time (HH:MM)")

                try:
                    # Convert all HTML time values to AM/PM format used in helpers
                    settings_payload = {
                        "morning_in_start": parse_html_time(request.form.get("morning_in_start", ""), "Morning start time"),
                        "morning_in_late": parse_html_time(request.form.get("morning_in_late", ""), "Morning late-after time"),
                        "morning_in_window_end": parse_html_time(request.form.get("morning_in_window_end", ""), "Morning window end"),
                        "lunch_out_start": parse_html_time(request.form.get("lunch_out_start", ""), "Lunch out start"),
                        "lunch_out_end": parse_html_time(request.form.get("lunch_out_end", ""), "Lunch out end"),
                        "afternoon_in_start": parse_html_time(request.form.get("afternoon_in_start", ""), "Afternoon start time"),
                        "afternoon_in_late": parse_html_time(request.form.get("afternoon_in_late", ""), "Afternoon late-after time"),
                        "afternoon_in_window_end": parse_html_time(request.form.get("afternoon_in_window_end", ""), "Afternoon window end"),
                        "time_out_start": parse_html_time(request.form.get("time_out_start", ""), "Time-out start"),
                    }

                    # Upsert settings into the settings table
                    for key, value in settings_payload.items():
                        cur.execute(
                            """
                            INSERT INTO settings (setting_key, setting_value)
                            VALUES (?, ?)
                            ON CONFLICT(setting_key) DO UPDATE SET setting_value=excluded.setting_value
                            """,
                            (key, value),
                        )

                    db.commit()
                    # Clear cache so new values are used immediately
                    clear_time_settings_cache()

                    # Reload settings for display
                    time_settings_raw = get_time_settings()
                    time_settings = {}
                    for key, value in time_settings_raw.items():
                        time_settings[key] = convert_time_for_html(value)

                    return render_template(
                        "admin/settings.html",
                        error=None,
                        success="Time settings saved successfully.",
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                except ValidationError as e:
                    logger.warning(f"Validation error in time settings: {str(e)}")
                    db.rollback()
                    # Reload settings for display even on validation error
                    time_settings_raw = get_time_settings()
                    time_settings = {}
                    for key, value in time_settings_raw.items():
                        time_settings[key] = convert_time_for_html(value)
                    return render_template(
                        "admin/settings.html",
                        error=str(e),
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
            
            if action == "update_profile":
                admin_id = session.get("admin_id")
                if not admin_id:
                    return render_template(
                        "admin/settings.html",
                        error="Session expired. Please log in again to update your profile.",
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                from config import ADMIN_PHOTOS_DIR, ALLOWED_IMAGE_EXTENSIONS, MAX_UPLOAD_SIZE
                display_name = (request.form.get("admin_display_name") or "").strip()
                if not display_name:
                    return render_template(
                        "admin/settings.html",
                        error="Display name is required.",
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                photo_path_value = current_admin.get("photo_path") if current_admin else ""
                file = request.files.get("admin_photo")
                if file and file.filename:
                    ext = (file.filename.rsplit(".", 1)[-1] or "").lower()
                    if ext not in ALLOWED_IMAGE_EXTENSIONS:
                        return render_template(
                            "admin/settings.html",
                            error="Invalid photo format. Allowed: " + ", ".join(sorted(ALLOWED_IMAGE_EXTENSIONS)),
                            success=None,
                            time_settings=time_settings,
                            current_admin=current_admin,
                        )
                    file.seek(0, 2)
                    size = file.tell()
                    file.seek(0)
                    if size > MAX_UPLOAD_SIZE:
                        return render_template(
                            "admin/settings.html",
                            error="Photo is too large. Maximum size is 10 MB.",
                            success=None,
                            time_settings=time_settings,
                            current_admin=current_admin,
                        )
                    os.makedirs(ADMIN_PHOTOS_DIR, exist_ok=True)
                    safe_name = secure_filename(f"admin_{admin_id}.{ext}")
                    filepath = os.path.join(ADMIN_PHOTOS_DIR, safe_name)
                    file.save(filepath)
                    photo_path_value = os.path.join("uploads", "admins", safe_name).replace("\\", "/")
                cur.execute("UPDATE admin SET name = ?, photo_path = ? WHERE id = ?", (display_name, photo_path_value or None, admin_id))
                db.commit()
                session["admin_name"] = display_name
                session["admin_photo_path"] = photo_path_value or ""
                logger.info(f"Admin profile updated: id={admin_id}, name={display_name}")
                return render_template(
                    "admin/settings.html",
                    error=None,
                    success="Profile updated successfully. Your name and photo are updated.",
                    time_settings=time_settings,
                    current_admin={"id": admin_id, "name": display_name, "photo_path": photo_path_value},
                )

            if action == "update_logo":
                from config import LOGO_UPLOADS_DIR, ALLOWED_IMAGE_EXTENSIONS, MAX_UPLOAD_SIZE
                file = request.files.get("organization_logo")
                if not file or not file.filename:
                    return render_template(
                        "admin/settings.html",
                        error="Please select an image file for the organization logo.",
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                ext = (file.filename.rsplit(".", 1)[-1] or "").lower()
                if ext not in ALLOWED_IMAGE_EXTENSIONS:
                    return render_template(
                        "admin/settings.html",
                        error="Invalid logo format. Allowed: " + ", ".join(sorted(ALLOWED_IMAGE_EXTENSIONS)),
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                file.seek(0, 2)
                size = file.tell()
                file.seek(0)
                if size > MAX_UPLOAD_SIZE:
                    return render_template(
                        "admin/settings.html",
                        error="Logo is too large. Maximum size is 10 MB.",
                        success=None,
                        time_settings=time_settings,
                        current_admin=current_admin,
                    )
                os.makedirs(LOGO_UPLOADS_DIR, exist_ok=True)
                safe_name = "logo." + ext
                filepath = os.path.join(LOGO_UPLOADS_DIR, safe_name)
                file.save(filepath)
                logo_static_path = os.path.join("uploads", "logo", safe_name).replace("\\", "/")
                cur.execute(
                    """
                    INSERT INTO settings (setting_key, setting_value) VALUES (?, ?)
                    ON CONFLICT(setting_key) DO UPDATE SET setting_value=excluded.setting_value
                    """,
                    ("organization_logo", logo_static_path),
                )
                db.commit()
                logger.info("Organization logo updated")
                return render_template(
                    "admin/settings.html",
                    error=None,
                    success="Organization logo updated successfully.",
                    time_settings=time_settings,
                    current_admin=current_admin,
                )

            if action == "mark_holiday":
                holiday_date = validate_required(request.form.get("holiday_date", ""), "Holiday Date")
                reason = sanitize_string(request.form.get("reason", "Holiday"), max_length=200) or "Holiday"
                
                cur.execute("SELECT id FROM employees WHERE status='Active'")
                employees = cur.fetchall()
                
                if not employees:
                    return render_template("admin/settings.html",
                                         error="No active employees found",
                                         success=None,
                                         time_settings=time_settings,
                                         current_admin=current_admin)
                
                marked_count = 0
                for emp in employees:
                    employee_id = emp["id"]
                    
                    cur.execute("""
                        SELECT attendance_id FROM attendance 
                        WHERE employee_id=? AND date=?
                    """, (employee_id, holiday_date))
                    existing = cur.fetchone()
                    
                    if existing:
                        cur.execute("""
                            UPDATE attendance 
                            SET morning_in='08:00 AM', 
                                lunch_out='12:00 PM',
                                afternoon_in='01:00 PM',
                                time_out='05:00 PM',
                                attendance_status='Present',
                                verification_method=?
                            WHERE employee_id=? AND date=?
                        """, (f"Admin: {reason}", employee_id, holiday_date))
                    else:
                        cur.execute("""
                            INSERT INTO attendance 
                            (employee_id, date, morning_in, lunch_out, afternoon_in, 
                             time_out, attendance_status, verification_method)
                            VALUES (?, ?, '08:00 AM', '12:00 PM', '01:00 PM', 
                                    '05:00 PM', 'Present', ?)
                        """, (employee_id, holiday_date, f"Admin: {reason}"))
                    
                    marked_count += 1
                
                db.commit()
                logger.info(f"Holiday marked for {marked_count} employees on {holiday_date}")
                return render_template("admin/settings.html",
                                     success=f"Successfully marked {marked_count} employees as present for {holiday_date} ({reason})",
                                     error=None,
                                     time_settings=time_settings,
                                     current_admin=current_admin)
            
            elif action == "mark_suspension":
                suspension_date = validate_required(request.form.get("suspension_date", ""), "Suspension Date")
                reason = sanitize_string(request.form.get("reason", "Suspension"), max_length=200) or "Suspension"
                
                cur.execute("SELECT id FROM employees WHERE status='Active'")
                employees = cur.fetchall()
                
                if not employees:
                    return render_template("admin/settings.html",
                                         error="No active employees found",
                                         success=None,
                                         time_settings=time_settings,
                                         current_admin=current_admin)
                
                marked_count = 0
                for emp in employees:
                    employee_id = emp["id"]
                    
                    cur.execute("""
                        SELECT attendance_id FROM attendance 
                        WHERE employee_id=? AND date=?
                    """, (employee_id, suspension_date))
                    existing = cur.fetchone()
                    
                    if existing:
                        cur.execute("""
                            UPDATE attendance 
                            SET morning_in='08:00 AM', 
                                lunch_out='12:00 PM',
                                afternoon_in='01:00 PM',
                                time_out='05:00 PM',
                                attendance_status='Present',
                                verification_method=?
                            WHERE employee_id=? AND date=?
                        """, (f"Admin: {reason}", employee_id, suspension_date))
                    else:
                        cur.execute("""
                            INSERT INTO attendance 
                            (employee_id, date, morning_in, lunch_out, afternoon_in, 
                             time_out, attendance_status, verification_method)
                            VALUES (?, ?, '08:00 AM', '12:00 PM', '01:00 PM', 
                                    '05:00 PM', 'Present', ?)
                        """, (employee_id, suspension_date, f"Admin: {reason}"))
                    
                    marked_count += 1
                
                db.commit()
                logger.info(f"Suspension marked for {marked_count} employees on {suspension_date}")
                return render_template("admin/settings.html",
                                     success=f"Successfully marked {marked_count} employees as present for {suspension_date} ({reason})",
                                     error=None,
                                     time_settings=time_settings,
                                     current_admin=current_admin)
        
        return render_template("admin/settings.html", error=None, success=None, time_settings=time_settings, current_admin=current_admin)
        
    except ValidationError as e:
        logger.warning(f"Validation error in settings: {str(e)}")
        # Get time settings even on validation error
        try:
            from utils.helpers import get_time_settings
            from datetime import datetime
            time_settings_raw = get_time_settings()
            def convert_time_for_html(time_str):
                try:
                    if time_str and ('AM' in time_str or 'PM' in time_str):
                        time_obj = datetime.strptime(time_str, "%I:%M %p")
                        return time_obj.strftime("%H:%M")
                    return time_str or ""
                except:
                    return time_str or ""
            time_settings = {k: convert_time_for_html(v) for k, v in time_settings_raw.items()}
        except:
            time_settings = {}
        current_admin = None
        try:
            if session.get("admin_id") and db is not None:
                cur = db.cursor()
                cur.execute("SELECT id, name, photo_path FROM admin WHERE id = ?", (session["admin_id"],))
                row = cur.fetchone()
                if row:
                    current_admin = {"id": row["id"], "name": row["name"] or "", "photo_path": (row["photo_path"] if "photo_path" in row.keys() else None) or ""}
        except Exception:
            pass
        return render_template("admin/settings.html", error=str(e), success=None, time_settings=time_settings, current_admin=current_admin)
    except Exception as e:
        # Only rollback if db was created
        try:
            if db is not None:
                db.rollback()
        except:
            pass
        logger.error(f"Error in settings: {str(e)}", exc_info=True)
        # Get time settings even on error
        try:
            from utils.helpers import get_time_settings
            from datetime import datetime
            time_settings_raw = get_time_settings()
            def convert_time_for_html(time_str):
                try:
                    if time_str and ('AM' in time_str or 'PM' in time_str):
                        time_obj = datetime.strptime(time_str, "%I:%M %p")
                        return time_obj.strftime("%H:%M")
                    return time_str or ""
                except:
                    return time_str or ""
            time_settings = {k: convert_time_for_html(v) for k, v in time_settings_raw.items()}
        except:
            time_settings = {}
        current_admin = None
        try:
            if session.get("admin_id") and db is not None:
                cur = db.cursor()
                cur.execute("SELECT id, name, photo_path FROM admin WHERE id = ?", (session["admin_id"],))
                row = cur.fetchone()
                if row:
                    current_admin = {"id": row["id"], "name": row["name"] or "", "photo_path": (row["photo_path"] if "photo_path" in row.keys() else None) or ""}
        except Exception:
            pass
        return render_template("admin/settings.html",
                             error=f"Error: {str(e)}",
                             success=None,
                             time_settings=time_settings,
                             current_admin=current_admin)


@admin_bp.route("/capture-face", methods=["POST"])
@admin_required
def capture_face():
    """Endpoint to capture and encode a single face image"""
    try:
        data = request.get_json()
        image_data = data.get("image")
        employee_id = data.get("employee_id")

        if not image_data:
            return jsonify({"success": False, "error": "No image data provided"}), 400

        encoding = face_utils.encode_face_from_base64(image_data)
        
        if encoding is None:
            return jsonify({"success": False, "error": "No face detected in image"}), 400

        if employee_id:
            face_utils.save_face(employee_id, encoding)
            logger.info(f"Face captured and saved for employee {employee_id}")

        return jsonify({"success": True, "message": "Face captured successfully"})
    except Exception as e:
        logger.error(f"Error capturing face: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500



# ==================== ADMIN ACCOUNT MANAGEMENT ====================

@admin_bp.route("/admins")
@admin_required
def manage_admins():
    """Admin account management page"""
    try:
        from db import get_db
        from utils.db_helpers import execute_query_safe
        db = get_db()
        
        # Get all admin accounts
        admins = execute_query_safe(
            db,
            "SELECT id, username, name FROM admin ORDER BY id",
            fetch_all=True
        )
        
        logger.info(f"Admin management page accessed by: {session.get('admin_name')}")
        return render_template("admin/admins.html", admins=admins or [])
    except Exception as e:
        logger.error(f"Error loading admin management page: {str(e)}", exc_info=True)
        return render_template("admin/admins.html", admins=[], error=str(e))


@admin_bp.route("/admins/create", methods=["POST"])
@admin_required
def create_admin():
    """Create a new admin account"""
    try:
        username = validate_required(request.form.get("username"), "Username")
        password = validate_required(request.form.get("password"), "Password")
        name = validate_required(request.form.get("name"), "Name")
        
        # Sanitize inputs
        username = sanitize_string(username, max_length=50)
        name = sanitize_string(name, max_length=100)
        
        from db import get_db
        from utils.db_helpers import execute_query_safe
        db = get_db()
        
        # Check if username already exists
        existing = execute_query_safe(
            db,
            "SELECT id FROM admin WHERE username=?",
            (username,),
            fetch_one=True
        )
        
        if existing:
            return jsonify({"success": False, "error": "Username already exists"}), 400
        
        # Hash password and create new admin
        hashed_pwd = hash_password(password)
        execute_query_safe(
            db,
            "INSERT INTO admin (username, password, name) VALUES (?, ?, ?)",
            (username, hashed_pwd, name)
        )
        
        db.commit()
        logger.info(f"New admin account created: {username} by {session.get('admin_name')}")
        log_admin_action(
            db,
            action="create_admin",
            details=f"Created admin '{username}'",
        )
        
        return jsonify({"success": True, "message": f"Admin account '{username}' created successfully"})
    except ValidationError as e:
        logger.warning(f"Admin creation validation error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 400
    except ValueError as e:
        # Typically raised by hash_password for weak/empty passwords
        logger.warning(f"Admin creation password error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 400
    except sqlite3.IntegrityError as e:
        logger.warning(f"Admin creation integrity error: {str(e)}")
        return jsonify({"success": False, "error": "Username already exists"}), 400
    except Exception as e:
        logger.error(f"Error creating admin: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": f"Error creating admin: {str(e)}"}), 500


@admin_bp.route("/admins/update/<int:admin_id>", methods=["POST"])
@admin_required
def update_admin(admin_id):
    """Update an admin account"""
    try:
        username = validate_required(request.form.get("username"), "Username")
        password = request.form.get("password")  # Optional - only update if provided
        name = validate_required(request.form.get("name"), "Name")
        
        # Sanitize inputs
        username = sanitize_string(username, max_length=50)
        name = sanitize_string(name, max_length=100)
        
        from db import get_db
        from utils.db_helpers import execute_query_safe
        db = get_db()
        
        # Check if admin exists
        existing = execute_query_safe(
            db,
            "SELECT id, username FROM admin WHERE id=?",
            (admin_id,),
            fetch_one=True
        )
        
        if not existing:
            return jsonify({"success": False, "error": "Admin account not found"}), 404
        
        # Check if username already exists (excluding current admin)
        username_check = execute_query_safe(
            db,
            "SELECT id FROM admin WHERE username=? AND id!=?",
            (username, admin_id),
            fetch_one=True
        )
        
        if username_check:
            return jsonify({"success": False, "error": "Username already exists"}), 400
        
        # Update admin (password only if provided)
        if password and password.strip():
            # New password provided - hash it
            execute_query_safe(
                db,
                "UPDATE admin SET username=?, password=?, name=? WHERE id=?",
                (username, hash_password(password), name, admin_id)
            )
        else:
            execute_query_safe(
                db,
                "UPDATE admin SET username=?, name=? WHERE id=?",
                (username, name, admin_id)
            )
        
        db.commit()
        logger.info(f"Admin account updated: ID {admin_id} by {session.get('admin_name')}")
        log_admin_action(
            db,
            action="update_admin",
            details=f"Updated admin ID {admin_id} (username: {username})",
        )
        
        return jsonify({"success": True, "message": f"Admin account '{username}' updated successfully"})
    except ValidationError as e:
        logger.warning(f"Admin update validation error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 400
    except ValueError as e:
        # Typically raised by hash_password for weak/empty passwords
        logger.warning(f"Admin update password error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 400
    except sqlite3.IntegrityError as e:
        logger.warning(f"Admin update integrity error: {str(e)}")
        return jsonify({"success": False, "error": "Username already exists"}), 400
    except Exception as e:
        logger.error(f"Error updating admin: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": f"Error updating admin: {str(e)}"}), 500


@admin_bp.route("/admins/delete/<int:admin_id>", methods=["POST"])
@admin_required
def delete_admin(admin_id):
    """Delete an admin account"""
    try:
        from db import get_db
        from utils.db_helpers import execute_query_safe
        db = get_db()
        
        # Get admin info before deletion for logging
        admin_info = execute_query_safe(
            db,
            "SELECT username FROM admin WHERE id=?",
            (admin_id,),
            fetch_one=True
        )
        
        if not admin_info:
            return jsonify({"success": False, "error": "Admin account not found"}), 404
        
        # Check if this is the only admin (prevent deleting last admin)
        admin_count = execute_query_safe(
            db,
            "SELECT COUNT(*) as count FROM admin",
            fetch_one=True
        )
        
        # sqlite3.Row does not support .get(); use index or key access
        total_admins = 0
        if admin_count is not None:
            try:
                total_admins = admin_count["count"]
            except (KeyError, TypeError):
                total_admins = admin_count[0]
        
        if total_admins <= 1:
            return jsonify({"success": False, "error": "Cannot delete the last admin account"}), 400
        
        # Delete admin
        execute_query_safe(
            db,
            "DELETE FROM admin WHERE id=?",
            (admin_id,)
        )
        
        db.commit()
        logger.info(f"Admin account deleted: ID {admin_id} ({admin_info['username']}) by {session.get('admin_name')}")
        log_admin_action(
            db,
            action="delete_admin",
            details=f"Deleted admin ID {admin_id} ({admin_info['username']})",
        )

        return jsonify({"success": True, "message": f"Admin account '{admin_info['username']}' deleted successfully"})
    except Exception as e:
        logger.error(f"Error deleting admin: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": f"Error deleting admin: {str(e)}"}), 500


@admin_bp.route("/admins/get/<int:admin_id>")
@admin_required
def get_admin(admin_id):
    """Get admin account details for editing"""
    try:
        from db import get_db
        from utils.db_helpers import execute_query_safe
        db = get_db()
        
        admin = execute_query_safe(
            db,
            "SELECT id, username, name FROM admin WHERE id=?",
            (admin_id,),
            fetch_one=True
        )
        
        if not admin:
            return jsonify({"success": False, "error": "Admin account not found"}), 404
        
        # Convert sqlite3.Row to dict
        admin_dict = {
            "id": admin["id"],
            "username": admin["username"] if "username" in admin.keys() else "",
            "name": admin["name"] if "name" in admin.keys() else ""
        }
        
        return jsonify({"success": True, "admin": admin_dict})
    except Exception as e:
        logger.error(f"Error getting admin: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


# ==================== EXPORT FUNCTIONALITY ====================

@admin_bp.route("/export/attendance/<format>")
@admin_required
def export_attendance(format):
    """Export attendance data in CSV, Excel, or PDF format"""
    try:
        from db import get_db
        from flask import Response, make_response
        from datetime import datetime
        import csv
        import io
        
        db = get_db()
        cur = db.cursor()
        
        # Get date filter if provided
        selected_date = request.args.get("date", "")
        
        # Build query
        if selected_date:
            query = """
                SELECT e.full_name, e.employee_code, e.department, a.date, 
                       a.morning_in, a.lunch_out, a.afternoon_in, a.time_out, 
                       a.attendance_status, a.verification_method
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                WHERE a.date=?
                ORDER BY e.full_name, a.date
            """
            params = (selected_date,)
        else:
            query = """
                SELECT e.full_name, e.employee_code, e.department, a.date, 
                       a.morning_in, a.lunch_out, a.afternoon_in, a.time_out, 
                       a.attendance_status, a.verification_method
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                ORDER BY a.date DESC, e.full_name
                LIMIT 1000
            """
            params = None
        
        cur.execute(query, params) if params else cur.execute(query)
        records = cur.fetchall()
        
        if format == 'csv':
            # CSV Export
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Employee Name', 'Employee Code', 'Department', 'Date',
                'Morning In', 'Lunch Out', 'Afternoon In', 'Time Out',
                'Status', 'Verification Method'
            ])
            
            # Write data
            for record in records:
                writer.writerow([
                    record["full_name"] or "",
                    record["employee_code"] or "",
                    record["department"] or "",
                    record["date"] or "",
                    record["morning_in"] or "",
                    record["lunch_out"] or "",
                    record["afternoon_in"] or "",
                    record["time_out"] or "",
                    record["attendance_status"] or "",
                    record["verification_method"] or ""
                ])
            
            output.seek(0)
            filename = f"attendance_{selected_date or 'all'}_{datetime.now().strftime('%Y%m%d')}.csv"
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif format == 'excel':
            # Excel Export
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance Records"
                
                # Header row with styling
                headers = [
                    'Employee Name', 'Employee Code', 'Department', 'Date',
                    'Morning In', 'Lunch Out', 'Afternoon In', 'Time Out',
                    'Status', 'Verification Method'
                ]
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Data rows
                for row_num, record in enumerate(records, 2):
                    ws.cell(row=row_num, column=1, value=record["full_name"] or "")
                    ws.cell(row=row_num, column=2, value=record["employee_code"] or "")
                    ws.cell(row=row_num, column=3, value=record["department"] or "")
                    ws.cell(row=row_num, column=4, value=record["date"] or "")
                    ws.cell(row=row_num, column=5, value=record["morning_in"] or "")
                    ws.cell(row=row_num, column=6, value=record["lunch_out"] or "")
                    ws.cell(row=row_num, column=7, value=record["afternoon_in"] or "")
                    ws.cell(row=row_num, column=8, value=record["time_out"] or "")
                    ws.cell(row=row_num, column=9, value=record["attendance_status"] or "")
                    ws.cell(row=row_num, column=10, value=record["verification_method"] or "")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"attendance_{selected_date or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError:
                logger.error("openpyxl not installed. Please install it: pip install openpyxl")
                return jsonify({"success": False, "error": "Excel export requires openpyxl library"}), 500
                
        elif format == 'pdf':
            # PDF Export
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                
                output = io.BytesIO()
                doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
                elements = []
                
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#1f2937'),
                    spaceAfter=20,
                    alignment=1  # Center
                )
                
                # Title
                title_text = f"Attendance Report - {selected_date or 'All Records'}"
                elements.append(Paragraph(title_text, title_style))
                elements.append(Spacer(1, 0.2*inch))
                
                # Prepare data
                data = [['Employee Name', 'Employee Code', 'Department', 'Date', 'Morning In', 
                        'Lunch Out', 'Afternoon In', 'Time Out', 'Status']]
                
                for record in records:
                    data.append([
                        record["full_name"] or "",
                        record["employee_code"] or "",
                        record["department"] or "",
                        record["date"] or "",
                        record["morning_in"] or "--",
                        record["lunch_out"] or "--",
                        record["afternoon_in"] or "--",
                        record["time_out"] or "--",
                        record["attendance_status"] or ""
                    ])
                
                # Create table
                table = Table(data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 
                                                0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                
                elements.append(table)
                
                # Build PDF
                doc.build(elements)
                output.seek(0)
                
                filename = f"attendance_{selected_date or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/pdf'
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError:
                logger.error("reportlab not installed. Please install it: pip install reportlab")
                return jsonify({"success": False, "error": "PDF export requires reportlab library"}), 500
        else:
            return jsonify({"success": False, "error": "Invalid format. Use csv, excel, or pdf"}), 400
            
    except Exception as e:
        logger.error(f"Error exporting attendance: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route("/export/employees/<format>")
@admin_required
def export_employees(format):
    """Export employees data in CSV, Excel, or PDF format"""
    try:
        from db import get_db
        from flask import Response, make_response
        from datetime import datetime
        import csv
        import io
        
        db = get_db()
        cur = db.cursor()
        
        # Get search filter if provided
        search_query = request.args.get("search", "").strip()
        
        # Build query
        if search_query:
            search_pattern = f"%{search_query}%"
            query = """
                SELECT full_name, employee_code, department, position, 
                       contact_number, email, employment_status, status
                FROM employees 
                WHERE full_name LIKE ? OR employee_code LIKE ? OR department LIKE ?
                ORDER BY full_name
            """
            params = (search_pattern, search_pattern, search_pattern)
        else:
            query = """
                SELECT full_name, employee_code, department, position, 
                       contact_number, email, employment_status, status
                FROM employees 
                ORDER BY full_name
            """
            params = None
        
        cur.execute(query, params) if params else cur.execute(query)
        employees = cur.fetchall()
        
        if format == 'csv':
            # CSV Export
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Full Name', 'Employee Code', 'Department', 'Position',
                'Contact Number', 'Email', 'Employment Status', 'Status'
            ])
            
            # Write data
            for emp in employees:
                writer.writerow([
                    emp["full_name"] or "",
                    emp["employee_code"] or "",
                    emp["department"] or "",
                    emp["position"] or "",
                    emp["contact_number"] or "",
                    emp["email"] or "",
                    emp["employment_status"] or "",
                    emp["status"] or "Active"
                ])
            
            output.seek(0)
            filename = f"employees_{datetime.now().strftime('%Y%m%d')}.csv"
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif format == 'excel':
            # Excel Export
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Employees"
                
                # Header row with styling
                headers = [
                    'Full Name', 'Employee Code', 'Department', 'Position',
                    'Contact Number', 'Email', 'Employment Status', 'Status'
                ]
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Data rows
                for row_num, emp in enumerate(employees, 2):
                    ws.cell(row=row_num, column=1, value=emp["full_name"] or "")
                    ws.cell(row=row_num, column=2, value=emp["employee_code"] or "")
                    ws.cell(row=row_num, column=3, value=emp["department"] or "")
                    ws.cell(row=row_num, column=4, value=emp["position"] or "")
                    ws.cell(row=row_num, column=5, value=emp["contact_number"] or "")
                    ws.cell(row=row_num, column=6, value=emp["email"] or "")
                    ws.cell(row=row_num, column=7, value=emp["employment_status"] or "")
                    ws.cell(row=row_num, column=8, value=emp["status"] or "Active")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"employees_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError:
                logger.error("openpyxl not installed")
                return jsonify({"success": False, "error": "Excel export requires openpyxl library"}), 500
                
        elif format == 'pdf':
            # PDF Export
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                
                output = io.BytesIO()
                doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
                elements = []
                
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#1f2937'),
                    spaceAfter=20,
                    alignment=1
                )
                
                # Title
                elements.append(Paragraph("Employee List", title_style))
                elements.append(Spacer(1, 0.2*inch))
                
                # Prepare data
                data = [['Full Name', 'Employee Code', 'Department', 'Position', 'Contact', 'Email', 'Status']]
                
                for emp in employees:
                    data.append([
                        emp["full_name"] or "",
                        emp["employee_code"] or "",
                        emp["department"] or "",
                        emp["position"] or "",
                        emp["contact_number"] or "",
                        emp["email"] or "",
                        emp["status"] or "Active"
                    ])
                
                # Create table
                table = Table(data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 1.5*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                
                elements.append(table)
                doc.build(elements)
                output.seek(0)
                
                filename = f"employees_{datetime.now().strftime('%Y%m%d')}.pdf"
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/pdf'
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except ImportError:
                logger.error("reportlab not installed")
                return jsonify({"success": False, "error": "PDF export requires reportlab library"}), 500
        else:
            return jsonify({"success": False, "error": "Invalid format. Use csv, excel, or pdf"}), 400
            
    except Exception as e:
        logger.error(f"Error exporting employees: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route("/export/dtr-pdf")
@admin_required
def export_dtr_pdf():
    """Export DTR (Daily Time Record) as PDF"""
    try:
        from db import get_db
        from flask import make_response, render_template
        from datetime import datetime
        import io
        
        employee_id = request.args.get("employee_id")
        month_year = request.args.get("month", "")
        
        if not employee_id or not month_year:
            return jsonify({"success": False, "error": "Employee ID and month are required"}), 400
        
        db = get_db()
        cur = db.cursor()
        
        # Get employee info
        cur.execute("SELECT * FROM employees WHERE id=?", (employee_id,))
        employee_info = cur.fetchone()
        
        if not employee_info:
            return jsonify({"success": False, "error": "Employee not found"}), 404
        
        # Get attendance records
        try:
            year, month = month_year.split("-")
            cur.execute("""
                SELECT date, morning_in, lunch_out, afternoon_in, time_out, attendance_status
                FROM attendance
                WHERE employee_id=? AND date LIKE ?
                ORDER BY date
            """, (employee_id, f"{year}-{month}-%"))
            
            attendance_records = cur.fetchall()
            
            dtr_data = {
                "year": year,
                "month": month,
                "records": {},
                "total_work_minutes": 0,
            }

            def _compute_work_minutes(morning_in, lunch_out, afternoon_in, time_out):
                """
                Compute total work minutes for the day.
                - Count morning segment only if BOTH morning_in AND lunch_out exist (complete morning half).
                - Count afternoon segment only if BOTH afternoon_in AND time_out exist (complete afternoon half).
                - Half-day is OK: only morning or only afternoon still gets totalled.
                - Do NOT total if a segment has only one punch (e.g. morning in but no lunch out).
                - Lunch break (12:00 PM - 1:00 PM) not counted; afternoon count starts at 1:00 PM if earlier.
                """
                try:
                    fmt = "%I:%M %p"
                    lunch_start = datetime.strptime("12:00 PM", fmt)
                    afternoon_sched_start = datetime.strptime("01:00 PM", fmt)
                    morning_minutes = 0
                    afternoon_minutes = 0

                    if morning_in and lunch_out:
                        m_in = datetime.strptime(morning_in, fmt)
                        l_out = datetime.strptime(lunch_out, fmt)
                        morning_end = min(l_out, lunch_start)
                        morning_minutes = max(
                            0, int((morning_end - m_in).total_seconds() / 60)
                        )

                    if afternoon_in and time_out:
                        a_in = datetime.strptime(afternoon_in, fmt)
                        t_out = datetime.strptime(time_out, fmt)
                        afternoon_start = max(a_in, afternoon_sched_start)
                        afternoon_minutes = max(
                            0, int((t_out - afternoon_start).total_seconds() / 60)
                        )

                    work_minutes = morning_minutes + afternoon_minutes
                    return work_minutes if work_minutes > 0 else 0
                except Exception:
                    return 0
            
            for record in attendance_records:
                day = record["date"].split("-")[2]  # keep zero-padded e.g. "01", "02"
                morning_in = record["morning_in"] or ""
                lunch_out = record["lunch_out"] or ""
                afternoon_in = record["afternoon_in"] or ""
                time_out = record["time_out"] or ""
                work_minutes = _compute_work_minutes(
                    morning_in, lunch_out, afternoon_in, time_out
                )
                dtr_data["records"][day] = {
                    "morning_in": morning_in,
                    "lunch_out": lunch_out,
                    "afternoon_in": afternoon_in,
                    "time_out": time_out,
                    "status": record["attendance_status"] or "",
                    "work_minutes": work_minutes,
                }
                dtr_data["total_work_minutes"] += work_minutes
        except Exception as e:
            logger.warning(f"Error processing DTR data: {str(e)}")
            dtr_data = {"year": year, "month": month, "records": {}}
        
        # Generate PDF using reportlab - two side-by-side DTR forms via Frames
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph, Spacer, Table, TableStyle, NextFrameFlowable
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
            from calendar import month_name
            
            output = io.BytesIO()
            page_w, page_h = letter
            margin = 0.35 * inch
            col_width = (page_w - 2 * margin - 0.5 * inch) / 2  # two cols with gap
            frame_height = page_h - 2 * margin
            
            # Left and right frames for two columns
            frame_left = Frame(margin, margin, col_width, frame_height, id='col1')
            frame_right = Frame(margin + col_width + 0.5*inch, margin, col_width, frame_height, id='col2')
            doc = BaseDocTemplate(output, pagesize=letter, leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0)
            doc.addPageTemplates([PageTemplate(id='TwoCol', frames=[frame_left, frame_right])])
            
            def build_one_dtr():
                flowables = []
                styles = getSampleStyleSheet()
                form_style = ParagraphStyle(
                    'DTRFormNo', parent=styles['Normal'],
                    fontSize=8, textColor=colors.black, fontName='Times-Roman',
                    alignment=TA_LEFT, spaceAfter=4
                )
                flowables.append(Paragraph("Civil Service Form No. 48", form_style))
                title_style = ParagraphStyle(
                    'DTRTitle', parent=styles['Normal'],
                    fontSize=14, textColor=colors.black, fontName='Times-Bold',
                    alignment=TA_CENTER, spaceAfter=4
                )
                flowables.append(Paragraph("DAILY TIME RECORD", title_style))
                o0o_style = ParagraphStyle('DTRo0o', parent=form_style, alignment=TA_CENTER)
                flowables.append(Paragraph("-----o0o-----", o0o_style))
                flowables.append(Spacer(1, 0.12*inch))
                name_style = ParagraphStyle(
                    'DTRName', parent=styles['Normal'],
                    fontSize=11, textColor=colors.black, fontName='Times-Bold',
                    alignment=TA_CENTER, spaceAfter=2
                )
                flowables.append(Paragraph(f"{employee_info['full_name'] or 'N/A'}", name_style))
                flowables.append(Paragraph("(Name)", form_style))
                flowables.append(Spacer(1, 0.1*inch))
                month_name_str = month_name[int(month)]
                month_style = ParagraphStyle(
                    'DTRMonth', parent=styles['Normal'],
                    fontSize=10, textColor=colors.black, fontName='Times-Roman',
                    alignment=TA_CENTER, spaceAfter=8
                )
                flowables.append(Paragraph(f"For the month of {month_name_str} {year}", month_style))
                flowables.append(Spacer(1, 0.08*inch))
                official_style = ParagraphStyle(
                    'DTROfficial', parent=styles['Normal'],
                    fontSize=8, textColor=colors.black, fontName='Times-Roman',
                    alignment=TA_LEFT, spaceAfter=8
                )
                flowables.append(Paragraph("Official hours for arrival and departure &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Regular days ____________ &nbsp;&nbsp; Saturdays ____________", official_style))
                flowables.append(Spacer(1, 0.08*inch))
                
                days_in_month = 31
                try:
                    from calendar import monthrange
                    days_in_month = monthrange(int(year), int(month))[1]
                except Exception:
                    pass
                table_data = [
                    ['Day', 'A.M.', 'A.M.', 'P.M.', 'P.M.', 'Undertime', 'Undertime'],
                    ['', 'Arrival', 'Departure', 'Arrival', 'Departure', 'Hours', 'Minutes']
                ]
                for day in range(1, days_in_month + 1):
                    day_str = str(day).zfill(2)
                    day_key = day_str
                    rec = dtr_data.get("records", {}).get(day_key)
                    ut_minutes = rec.get("work_minutes", 0) if rec else 0
                    if rec:
                        has_full_times = bool(
                            rec.get("morning_in")
                            and rec.get("lunch_out")
                            and rec.get("afternoon_in")
                            and rec.get("time_out")
                        )
                        if ut_minutes:
                            ut_hours_val = ut_minutes // 60
                            ut_mins_val = ut_minutes % 60
                        elif has_full_times:
                            ut_hours_val = 0
                            ut_mins_val = 0
                        else:
                            ut_hours_val = ""
                            ut_mins_val = ""

                        table_data.append([
                            str(day),
                            rec.get("morning_in", ""),
                            rec.get("lunch_out", ""),
                            rec.get("afternoon_in", ""),
                            rec.get("time_out", ""),
                            ut_hours_val,
                            ut_mins_val,
                        ])
                    else:
                        table_data.append([str(day), "", "", "", "", "", ""])

                total_ut = dtr_data.get("total_work_minutes", 0) or 0
                table_data.append([
                    "Total", "", "", "", "",
                    total_ut // 60,
                    total_ut % 60,
                ])
                col_widths = [0.32*inch, 0.58*inch, 0.58*inch, 0.58*inch, 0.58*inch, 0.38*inch, 0.38*inch]
                tbl = Table(table_data, colWidths=col_widths)
                tbl.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, 1), 'Times-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 1), 7),
                    ('ALIGN', (0, 0), (0, -1), TA_LEFT),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#e5e5e5')),
                    ('TEXTCOLOR', (0, 0), (-1, 1), colors.black),
                    ('FONTNAME', (0, 2), (-1, -2), 'Times-Roman'),
                    ('FONTSIZE', (0, 2), (-1, -2), 7),
                    ('FONTNAME', (0, -1), (-1, -1), 'Times-Bold'),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ]))
                flowables.append(tbl)
                flowables.append(Spacer(1, 0.2*inch))
                cert_style = ParagraphStyle(
                    'DTRCert', parent=styles['Normal'],
                    fontSize=8, textColor=colors.black, fontName='Times-Italic',
                    alignment=TA_JUSTIFY, spaceAfter=12
                )
                flowables.append(Paragraph("I certify on my honor that the above is a true and correct report of the hours of work performed, record of which was made daily at the time of arrival and departure from office.", cert_style))
                flowables.append(Paragraph("_________________________________________________________________________", form_style))
                flowables.append(Paragraph("(Signature)", form_style))
                flowables.append(Spacer(1, 0.15*inch))
                flowables.append(Paragraph("VERIFIED as to the prescribed office hours:", cert_style))
                flowables.append(Spacer(1, 0.08*inch))
                flowables.append(Paragraph("_________________________________________________________________________", form_style))
                flowables.append(Paragraph("In Charge", ParagraphStyle('DTRInCharge', parent=form_style, fontName='Times-Bold', alignment=TA_CENTER)))
                return flowables
            
            # Build story: first column content, then force next frame, then second column (same)
            story = []
            story.extend(build_one_dtr())
            story.append(NextFrameFlowable())
            story.extend(build_one_dtr())
            doc.build(story)
            output.seek(0)
            
            filename = f"DTR_{employee_info['employee_code'] or employee_id}_{month_year}.pdf"
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            logger.error("reportlab not installed")
            return jsonify({"success": False, "error": "PDF export requires reportlab library"}), 500
        except Exception as e:
            logger.error(f"Error generating DTR PDF: {str(e)}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
            
    except Exception as e:
        logger.error(f"Error exporting DTR: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500